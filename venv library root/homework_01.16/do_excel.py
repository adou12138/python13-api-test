# 当前项目的名称: python_13 
# 新文件名称：do_excel 
# 当前登录名：LuckyLu
# 创建日期：2019/1/18 14:52
# 文件IDE名称：PyCharm

from openpyxl import Workbook
from openpyxl import load_workbook

def create_excel(excel_name):  # 创建一个excel
    # 新建一个Excel
    wb = Workbook()  # 新建一个
    wb.save(excel_name)  # 必须要保存 否则不能创建成功

class Cases:
    '这个是一个存储数据的类'
    def __init__(self):  # 存储初始化数据
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None

class DoExcel:
    '这是一个excel测试数据的类'
    def __init__(self, excel_name, excel_sheet_name):
        self.excel_name = excel_name
        self.excel_sheet_name = excel_sheet_name
        # self.test_button = test_button

    def read_excel(self):  # 读取数据
        wb = load_workbook(self.excel_name)  # 定位excel
        sheet = wb[self.excel_sheet_name]  # 定为表单
        # res = sheet.cell(row, col).value  # 定位单元格

        case = []
        for i in range(2, sheet.max_row + 1):
            row_case = Cases()
            row_case.case_id = sheet.cell(row=i, column=1).value
            row_case.title = sheet.cell(row=i, column=2).value
            row_case.url = sheet.cell(row=i, column=3).value
            row_case.data = sheet.cell(row=i, column=4).value
            row_case.method = sheet.cell(row=i, column=5).value
            row_case.expected = sheet.cell(row=i, column=6).value
            case.append(row_case)
        return case

    def write_excel(self, row, col, value):  # 写入数据
        wb = load_workbook(self.excel_name)  # 定位excel
        sheet = wb[self.excel_sheet_name]  # 定为表单

        sheet.cell(row, col).value = value  # 写入数据
        wb.save(self.excel_name)  # 保存数据

if __name__ == '__main__':
    # create_excel('luckytest.xlsx')
    cases = DoExcel('../datas/luckytest.xlsx', 'login').read_excel()
    print(cases)
    print(type(cases))
    # write = DoExcel('../datas/luckytest.xlsx', 'login').write_excel(2, 8, 'pass')
    # print(write)

