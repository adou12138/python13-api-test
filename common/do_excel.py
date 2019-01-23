# 当前项目的名称: python_13 
# 新文件名称：do_excel 
# 当前登录名：LuckyLu
# 创建日期：2019/1/11 14:52
# 文件IDE名称：PyCharm

from openpyxl import Workbook
from openpyxl import load_workbook

from common import contants  # 引用路径地址
from common.test_api_config import ReadConfig  # 导入url配置

data_max_mobilephone = eval(ReadConfig(contants.test_api_conf_file).get_value('MaxMobilePhone', 'mobilephone'))  # 数据库最大电话号码


# from week_7.class_unittest_test.test_config import ReadConfig
# test_button = ReadConfig('test_case.conf').get_value('CASE', 'button')
'''
params=datas??
'''

def create_excel(excel_name):  # 创建一个excel
    # 新建一个Excel
    wb = Workbook()  # 新建一个
    wb.save(excel_name)  # 必须要保存 否则不能创建成功

class Cases:
    '这个是一个存储数据的类'
    def __init__(self):  # 存储初始化数据
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None

class DoExcel:
    '这是一个excel测试数据的类'
    excel_file_name = None

    def __init__(self, excel_file_name, excel_sheet_name):  #, test_button):  # 定义初始化函数
        try:
            # 操作的文件
            self.excel_file_name = excel_file_name
            # 实例化一个workbook对象
            self.workbook = load_workbook(filename=self.excel_file_name)
            # 异常处理
        except FileNotFoundError as e:
            # 文件未找到异常处理
            print('{0} not found,please check file path', format(excel_file_name))
            raise e
        self.excel_sheet_name = excel_sheet_name
        # self.test_button = test_button

    def read_excel(self):  # 读取数据
        sheet = self.workbook[self.excel_sheet_name]  # 定为表单

        case = []
        for i in range(2, sheet.max_row + 1):
            row_case = Cases()
            row_case.case_id = sheet.cell(row=i, column=1).value
            row_case.title = sheet.cell(row=i, column=2).value

            row_case.url = sheet.cell(row=i, column=3).value  # 没有配置url
            # print(row_case.url)
            # row_case.url = ReadConfig(contants.test_api_conf_file).get_value('URL', 'path_url') + str(sheet.cell(row=i, column=3).value)  # 配置url方式
            # print(row_case.url)
            # print(type(row_case.url))

            row_case.data = sheet.cell(row=i, column=4).value
            # print(type(row_case.data))  # 字符串

            """配置文件-手机号码最大
            import json
            data_old = eval(sheet.cell(row=i, column=4).value)
            # mobilephone = data_old['mobilephone']
            # print(mobilephone)
            # print(type(mobilephone))
            mobilephone = data_max_mobilephone+1
            # print(mobilephone)
            # print(type(mobilephone))
            data_old['mobilephone'] = str(mobilephone)
            row_case.data = json.dumps(data_old)
            # print(row_case.data)
            # print(type(row_case.data))
            """

            row_case.method = sheet.cell(row=i, column=5).value
            row_case.expected = sheet.cell(row=i, column=6).value
            # if type(case.expected) == int:
            #     case.expected = str(case.expected)
            case.append(row_case)  # 将case放到cases 列表里面
        return case

        # if test_button == 'all':
        #     case = []
        #     for i in range(2, sheet.max_row+1):
        #         row_case = Cases()
        #         row_case.case_id = sheet.cell(row=i, column=1).value
        #         row_case.title = sheet.cell(row=i, column=2).value
        #         row_case.a = sheet.cell(row=i, column=3).value
        #         row_case.b = sheet.cell(row=i, column=4).value
        #         row_case.expected = sheet.cell(row=i, column=5).value
        #         case.append(row_case)
        #         # print(row_case)
        #     # print(case)
        # else:
        #     case = []
        #     for i in eval(test_button):
        #         row_case = Cases()
        #         row_case.case_id = sheet.cell(row=i+1, column=1).value
        #         row_case.title = sheet.cell(row=i+1, column=2).value
        #         row_case.a = sheet.cell(row=i+1, column=3).value
        #         row_case.b = sheet.cell(row=i+1, column=4).value
        #         row_case.expected = sheet.cell(row=i+1, column=5).value
        #         case.append(row_case)
        # return case

    def write_excel(self, row, actual, result):  # 写入数据
        sheet = self.workbook[self.excel_sheet_name]  # 获取sheet
        sheet.cell(row, 7).value = actual  # 写入实际测试结果
        sheet.cell(row, 8).value = result  # 写入执行结果
        self.workbook.save(filename=self.excel_file_name)  # 保存数据



if __name__ == '__main__':
    # create_excel('luckytest.xlsx')
    # cases = DoExcel('luckytest.xlsx', 'add', test_button).read_excel()
    # print(cases)
    # write = DoExcel('luckytest.xlsx', 'sub').write_excel(2, 6, 'pass')
    # print(write)

    # cases = DoExcel('../datas/luckytest.xlsx', 'register').read_excel()
    # cases = DoExcel(contants.excel_file, 'register').read_excel()
    # print(cases)
    # print(type(cases))

    # 不需要转json，直接字符串写入就可以了
    # write = do_excel.write_excel(2, str({"mobilephone": "15777777777", "pwd": "", "regname": "luckytest"}), "True")
    # print(write)
    # print("*"*50)
    # do_excel = DoExcel(contants.excel_file)
    # cases_login = do_excel.read_excel("login")
    # print(cases_login)
    # write = do_excel.write_excel(2, str({"mobilephone": "15777777777", "pwd": "234", "regname": "luckytest"}),"False")
    # print(write)

    # cases_register = DoExcel(contants.excel_file, "register").read_excel()
    # print(cases_register)
    # DoExcel(contants.excel_file, "register").write_excel(2, str({"mobilephone": "15777777777", "pwd": "234", "regname": "luckytest"}),"False")
    # print("*"*50)
    # cases_login = DoExcel(contants.excel_file,"login").read_excel()
    # print(cases_login)
    cases_recharge = DoExcel(contants.excel_file, "recharge").read_excel()
    print(cases_recharge)
    # DoExcel(contants.excel_file, "recharge").write_excel(2, str({"status":1,"code":"10001","data":{"id":1114421,"regname":"小蜜蜂","pwd":"E10ADC3949BA59ABBE56E057F20F883E","mobilephone":"15666666666","leaveamount":"806838.00","type":"1","regtime":"2019-01-15 13:59:38.0"},"msg":"充值成功"}),"False")
    # print("*"*50)
    # cases_recharge = DoExcel(contants.excel_file, "withdraw").read_excel()
    # print(cases_recharge)
    # DoExcel(contants.excel_file, "recharge").write_excel(2, str({"status":1,"code":"10001","data":{"id":1114421,"regname":"小蜜蜂","pwd":"E10ADC3949BA59ABBE56E057F20F883E","mobilephone":"15666666666","leaveamount":"806838.00","type":"1","regtime":"2019-01-15 13:59:38.0"},"msg":"充值成功"}),"False")
    # print("*"*50)
    # cases_add = DoExcel(contants.excel_file, "add").read_excel()
    # print(cases_add)
    # DoExcel(contants.excel_file, "add").write_excel(2, str({"status":1,"code":"10001","data":{"id":1114421,"regname":"小蜜蜂","pwd":"E10ADC3949BA59ABBE56E057F20F883E","mobilephone":"15666666666","leaveamount":"806838.00","type":"1","regtime":"2019-01-15 13:59:38.0"},"msg":"充值成功"}),"False")
    # print("*"*50)
    cases_audit = DoExcel(contants.excel_file, "audit").read_excel()
    print(cases_audit)
    DoExcel(contants.excel_file, "audit").write_excel(2, str({"status":1,"code":"10001","data":{"id":1114421,"regname":"小蜜蜂","pwd":"E10ADC3949BA59ABBE56E057F20F883E","mobilephone":"15666666666","leaveamount":"806838.00","type":"1","regtime":"2019-01-15 13:59:38.0"},"msg":"充值成功"}),"False")
    print("*"*50)
